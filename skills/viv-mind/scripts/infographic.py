#!/usr/bin/env python3
"""Generate a detailed watercolor infographic for a journal entry via Venice AI."""

import argparse
import base64
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

API_BASE = "https://api.venice.ai/api/v1"
USER_AGENT = "VivMindSkill/1.0"
DEFAULT_MODEL = "flux-2-max"

DEFAULT_REPO = Path.home() / ".openclaw" / "workspace" / "viv-mind-scaffold"
DEFAULT_ASSETS = DEFAULT_REPO / "docs" / "assets" / "infographics"
DEFAULT_JOURNAL = DEFAULT_REPO / "data" / "viv-mind-journal.xlsx"


def get_api_key() -> str | None:
    """Get Venice API key from env or clawdbot config."""
    key = os.environ.get("VENICE_API_KEY", "").strip()
    if key:
        return key
    clawdbot = Path.home() / ".clawdbot" / "clawdbot.json"
    if clawdbot.exists():
        try:
            cfg = json.loads(clawdbot.read_text(encoding="utf-8"))
            for skill_name in ("venice-ai", "venice-ai-media", "viv-mind"):
                k = (
                    cfg.get("skills", {})
                    .get("entries", {})
                    .get(skill_name, {})
                    .get("env", {})
                    .get("VENICE_API_KEY", "")
                )
                if k:
                    return k.strip()
        except (json.JSONDecodeError, OSError):
            pass
    return None


def require_api_key() -> str:
    key = get_api_key()
    if not key:
        print("Error: VENICE_API_KEY not set", file=sys.stderr)
        sys.exit(2)
    return key


def load_entry(journal_path: Path, entry_id: str) -> dict[str, str]:
    """Load an entry from the Excel journal."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required")

    wb = openpyxl.load_workbook(journal_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("id") + 1
    except ValueError:
        raise RuntimeError("Journal missing id column")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == entry_id:
            return {h: (ws.cell(row=row, column=i + 1).value or "")
                    for i, h in enumerate(headers)}
    raise RuntimeError(f"Entry not found: {entry_id}")


def build_prompt(entry: dict[str, str]) -> str:
    """Construct a detailed watercolor infographic prompt."""
    title = entry.get("title", "Untitled")[:120]
    summary = entry.get("summary", "")[:300]
    author = entry.get("author", "")[:80]
    tags = entry.get("tags", "")
    source_type = entry.get("source_type", "article")
    source_url = entry.get("source_url", "")

    tag_text = tags if isinstance(tags, str) else "; ".join(str(t) for t in tags)
    tag_bullets = "\n".join(f"- {t.strip()}" for t in str(tag_text).split(";") if t.strip())[:300]

    prompt = f"""Create a detailed vertical watercolor infographic, portrait orientation 1024x1280.

Theme: whimsical, hand-painted, dreamy journal spread.
Color palette: magenta pink, cerulean blue, warm orange, soft cream paper texture.

Center title in elegant brush lettering:
"{title}"

Below the title, a soft watercolor illustration that captures the mood of this {source_type}:
{summary}

Add floating visual elements: watercolor splashes, doodled stars, organic shapes, delicate vines, subtle dot patterns.
Include a small byline area: "by {author or 'unknown'}".
Include a tag cloud area with these tags as tiny painted badges:
{tag_bullets or '- mind'}

At the bottom, include a tiny source ribbon:
Source: {source_url[:80]}

Style notes: no photorealism, no hard edges, painterly watercolor textures, high detail, joyful and contemplative mood, generous white space, magazine cover quality.
"""
    return prompt.strip()


def generate_image(
    api_key: str,
    prompt: str,
    width: int = 1024,
    height: int = 1280,
    model: str = DEFAULT_MODEL,
    timeout: int = 300,
) -> bytes:
    """Call Venice image generation and return image bytes.

    The Venice image generate endpoint can take 30-180s to return. We use a
    long timeout and rely on caller cancellation if needed.
    """
    url = f"{API_BASE}/image/generate"
    payload = {
        "model": model,
        "prompt": prompt,
        "width": width,
        "height": height,
        "format": "png",
        "safe_mode": False,
        "hide_watermark": True,
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": USER_AGENT,
        },
        data=body,
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Venice API error ({e.code}): {error_body}") from e

    images = data.get("images", [])
    if not images:
        raise RuntimeError("No image returned from Venice API")
    return base64.b64decode(images[0])


def slugify(text: str, max_len: int = 40) -> str:
    import re
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return (text or "image")[:max_len]


def update_entry_image(journal_path: Path, entry_id: str, image_path: Path) -> None:
    """Update image_path column for the entry."""
    try:
        import openpyxl
    except ImportError:
        return
    wb = openpyxl.load_workbook(journal_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("id") + 1
        img_col = headers.index("image_path") + 1
    except ValueError:
        return
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == entry_id:
            ws.cell(row=row, column=img_col, value=str(image_path))
            wb.save(journal_path)
            return


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate watercolor infographic for a journal entry.")
    parser.add_argument("--id", required=True, help="Journal entry id")
    parser.add_argument("--journal", default=str(DEFAULT_JOURNAL), help="Path to journal xlsx")
    parser.add_argument("--assets", default=str(DEFAULT_ASSETS), help="Output directory for infographics")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"Venice image model (default: {DEFAULT_MODEL})")
    parser.add_argument("--width", type=int, default=1024)
    parser.add_argument("--height", type=int, default=1280)
    args = parser.parse_args()

    api_key = require_api_key()
    journal_path = Path(args.journal).expanduser()
    assets_dir = Path(args.assets).expanduser()
    assets_dir.mkdir(parents=True, exist_ok=True)

    entry = load_entry(journal_path, args.id)
    prompt = build_prompt(entry)

    print(f"Generating infographic for entry {args.id}...")
    print("This may take 30-180 seconds. Use Ctrl+C to cancel.")
    image_bytes = generate_image(api_key, prompt, args.width, args.height, args.model)

    filename = f"{args.id}-{slugify(entry.get('title', 'untitled'))}.png"
    image_path = assets_dir / filename
    image_path.write_bytes(image_bytes)
    print(f"Saved: {image_path} ({len(image_bytes) // 1024}KB)")

    update_entry_image(journal_path, args.id, image_path)
    print(f"Updated journal image_path for {args.id}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
