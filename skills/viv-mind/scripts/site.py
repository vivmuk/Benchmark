#!/usr/bin/env python3
"""Generate an Instagram-style HTML feed from the viv-mind journal."""

import argparse
import json
import re
from pathlib import Path
from urllib.parse import quote

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

DEFAULT_REPO = Path.home() / ".openclaw" / "workspace" / "viv-mind-scaffold"
DEFAULT_JOURNAL = DEFAULT_REPO / "data" / "viv-mind-journal.xlsx"
DEFAULT_OUTPUT = DEFAULT_REPO / "docs" / "index.html"


def load_entries(journal_path: Path) -> list[dict[str, str]]:
    """Load entries from xlsx, newest last."""
    if openpyxl is None:
        raise RuntimeError("openpyxl required")
    if not journal_path.exists():
        return []
    wb = openpyxl.load_workbook(journal_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    entries = []
    for row in range(2, ws.max_row + 1):
        entry = {}
        for i, h in enumerate(headers):
            entry[h] = ws.cell(row=row, column=i + 1).value or ""
        entries.append(entry)
    return entries


def _split_tags(tag_value: str) -> list[str]:
    if not tag_value:
        return []
    tags = re.split(r"[;,]", tag_value)
    return [t.strip() for t in tags if t.strip()]


def _asset_url(image_path: str) -> str:
    """Convert absolute path to relative URL served from docs/index.html."""
    if not image_path:
        return ""
    paths = [p.strip() for p in str(image_path).split(";") if p.strip()]
    result = []
    for raw in paths:
        p = Path(raw)
        if not p.exists():
            continue
        try:
            rel = p.relative_to(DEFAULT_REPO)
        except ValueError:
            rel = Path("assets") / "infographics" / p.name
        # index.html lives in docs/, so strip the leading docs/ prefix.
        rel_str = str(rel).replace("\\", "/")
        if rel_str.startswith("docs/"):
            rel_str = rel_str[5:]
        result.append(quote(rel_str, safe="/"))
    return result


def _format_date(date_value: str) -> str:
    """Normalize published_date to 'Mon D, YYYY' or 'Mon YYYY'."""
    if not date_value:
        return ""
    raw = str(date_value).strip()
    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", raw)
    if m:
        y, mo, d = m.groups()
        return f"{_month_name(mo)} {int(d)}, {y}"
    # YYYY-MM
    m = re.match(r"^(\d{4})-(\d{2})$", raw)
    if m:
        y, mo = m.groups()
        return f"{_month_name(mo)} {y}"
    # YYYYMMDD
    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", raw)
    if m:
        y, mo, d = m.groups()
        return f"{_month_name(mo)} {int(d)}, {y}"
    # YYYY
    m = re.match(r"^(\d{4})$", raw)
    if m:
        return raw
    return raw


def _month_name(mo: str) -> str:
    names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
        "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
        "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
    }
    return names.get(mo, mo)


def _summary_html(summary: str, eid: str) -> str:
    """Render summary with expand/collapse toggle when long. summary is already HTML-escaped."""
    if not summary:
        return '<p></p>'
    text = summary
    if len(text) <= 280:
        return f'<p>{text}</p>'
    short = text[:280].rsplit(" ", 1)[0]
    return f'''<p class="summary-short" id="summary-short-{eid}">{short}… <button class="summary-toggle" type="button" onclick="toggleSummary('{eid}')" aria-expanded="false">Read more</button></p>
        <p class="summary-full" id="summary-full-{eid}" style="display:none">{text} <button class="summary-toggle" type="button" onclick="toggleSummary('{eid}')" aria-expanded="true">Show less</button></p>'''


def escape_html(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def build_feed(entries: list[dict[str, str]]) -> str:
    """Render feed HTML."""
    cards = []
    for entry in reversed(entries):
        eid = escape_html(str(entry.get("id", "")))
        title = escape_html(str(entry.get("title", "Untitled")))
        source_url = escape_html(str(entry.get("source_url", "")))
        source_type = escape_html(str(entry.get("source_type", "link")))
        author = escape_html(str(entry.get("author", "")))
        summary = escape_html(str(entry.get("summary", "")))
        published = _format_date(str(entry.get("published_date", "")))
        tags = _split_tags(str(entry.get("tags", "")))
        image_paths = _asset_url(str(entry.get("image_path", "")))

        tag_html = "".join(
            f'<span class="tag">#{escape_html(tag)}</span>' for tag in tags
        )

        if len(image_paths) > 1:
            slides_html = "\n".join(
                f'<div class="carousel-slide"><img src="{p}" alt="{title} slide {i+1}" loading="lazy" onclick="openLightbox(\'{p}\', \'{title}\')" /></div>'
                for i, p in enumerate(image_paths)
            )
            dots_html = "".join(
                f'<span class="carousel-dot" data-index="{i}"></span>'
                for i in range(len(image_paths))
            )
            image_html = f"""<div class="carousel">
          <div class="carousel-track">{slides_html}</div>
          <div class="carousel-dots">{dots_html}</div>
        </div>"""
        elif image_paths:
            image_html = f'<img src="{image_paths[0]}" alt="{title}" loading="lazy" onclick="openLightbox(\'{image_paths[0]}\', \'{title}\')" />'
        else:
            image_html = f'<div class="placeholder">{source_type.upper()}</div>'

        cards.append(f"""
    <article class="card">
      <header class="card-header">
        <span class="avatar">🧠</span>
        <div class="meta">
          <div class="title">{title}</div>
          <div class="subtitle">{author or source_type} · {published}</div>
        </div>
      </header>
      <div class="media">
        {image_html}
      </div>
      <div class="caption">
        {_summary_html(summary, eid)}
        <div class="tags">{tag_html}</div>
      </div>
      <footer class="card-footer">
        <a class="source" href="{source_url}" target="_blank" rel="noopener">View source ↗</a>
        <span class="id">#{eid}</span>
      </footer>
    </article>
""")

    cards_html = "\n".join(cards)

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Viv Mind — Curated Feed</title>
  <style>
    :root {{
      --bg: #fff7f9;
      --card: #ffffff;
      --ink: #2a1f3d;
      --muted: #7b6b8d;
      --magenta: #d946ef;
      --blue: #3b82f6;
      --orange: #f97316;
      --radius: 24px;
      --shadow: 0 12px 30px rgba(42, 31, 61, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: radial-gradient(circle at top left, #ffe4f5 0%, var(--bg) 40%, #e0f2fe 100%);
      color: var(--ink);
      padding: 40px 16px;
    }}
    .container {{
      max-width: 720px;
      margin: 0 auto;
    }}
    header.hero {{
      text-align: center;
      margin-bottom: 40px;
    }}
    header.hero h1 {{
      font-size: 2.4rem;
      margin: 0 0 8px;
      letter-spacing: -0.03em;
    }}
    header.hero p {{
      color: var(--muted);
      margin: 0;
      font-size: 1.05rem;
    }}
    .feed {{
      display: grid;
      gap: 32px;
    }}
    .card {{
      background: var(--card);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow: hidden;
      border: 1px solid rgba(255, 255, 255, 0.8);
    }}
    .card-header {{
      display: flex;
      align-items: center;
      gap: 12px;
      padding: 16px 18px 10px;
    }}
    .avatar {{
      width: 42px;
      height: 42px;
      display: grid;
      place-items: center;
      font-size: 1.4rem;
      background: linear-gradient(135deg, var(--magenta), var(--orange));
      border-radius: 50%;
    }}
    .meta {{ flex: 1; }}
    .title {{ font-weight: 700; font-size: 1.05rem; line-height: 1.25; }}
    .subtitle {{ color: var(--muted); font-size: 0.82rem; margin-top: 2px; }}
    .media {{
      display: block;
      width: 100%;
      background: linear-gradient(135deg, #fce7f3, #dbeafe, #ffedd5);
      overflow: hidden;
    }}
    .media img {{
      width: 100%;
      height: auto;
      object-fit: contain;
      display: block;
    }}
    .carousel {{
      position: relative;
      width: 100%;
      overflow: hidden;
    }}
    .carousel-track {{
      display: flex;
      width: 100%;
      transition: transform 0.3s ease;
    }}
    .carousel-slide {{
      min-width: 100%;
    }}
    .carousel-slide img {{
      width: 100%;
      height: auto;
      object-fit: contain;
      display: block;
    }}
    .carousel-dots {{
      position: absolute;
      bottom: 12px;
      left: 50%;
      transform: translateX(-50%);
      display: flex;
      gap: 8px;
      z-index: 2;
    }}
    .carousel-dot {{
      width: 8px;
      height: 8px;
      border-radius: 50%;
      background: rgba(255,255,255,0.7);
      cursor: pointer;
      box-shadow: 0 1px 3px rgba(0,0,0,0.2);
    }}
    .carousel-dot.active {{ background: #d946ef; }}
    .placeholder {{
      width: 100%;
      height: 100%;
      display: grid;
      place-items: center;
      color: var(--muted);
      font-weight: 700;
      font-size: 1.2rem;
      letter-spacing: 0.1em;
    }}
    .caption {{
      padding: 16px 18px 8px;
      font-size: 0.96rem;
      line-height: 1.5;
    }}
    .caption p {{ margin: 0 0 12px; }}
    .summary-short {{ margin: 0 0 12px; }}
    .tags {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}
    .tag {{
      background: #f3e8ff;
      color: #7e22ce;
      padding: 4px 10px;
      border-radius: 999px;
      font-size: 0.78rem;
      font-weight: 600;
    }}
    .card-footer {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding: 10px 18px 16px;
      font-size: 0.85rem;
    }}
    .source {{
      color: var(--blue);
      text-decoration: none;
      font-weight: 600;
    }}
    .source:hover {{ text-decoration: underline; }}
    .id {{ color: var(--muted); font-family: ui-monospace, SFMono-Regular, monospace; }}
    .media {{ cursor: zoom-in; }}
    .lightbox {{
      position: fixed;
      inset: 0;
      background: rgba(0, 0, 0, 0.92);
      display: flex;
      align-items: center;
      justify-content: center;
      z-index: 1000;
      opacity: 0;
      pointer-events: none;
      transition: opacity 0.2s ease;
    }}
    .lightbox.active {{
      opacity: 1;
      pointer-events: auto;
    }}
    .lightbox img {{
      max-width: 96vw;
      max-height: 94vh;
      width: auto;
      height: auto;
      object-fit: contain;
      border-radius: 12px;
      box-shadow: 0 24px 60px rgba(0,0,0,0.4);
    }}
    .lightbox-close {{
      position: absolute;
      top: 16px;
      right: 20px;
      color: #fff;
      font-size: 2rem;
      background: none;
      border: none;
      cursor: pointer;
      line-height: 1;
    }}
    .lightbox-caption {{
      position: absolute;
      bottom: 16px;
      left: 16px;
      right: 16px;
      color: #fff;
      text-align: center;
      font-size: 1rem;
      text-shadow: 0 1px 4px rgba(0,0,0,0.6);
      pointer-events: none;
    }}
    .summary-toggle {{
      background: none;
      border: none;
      color: var(--blue);
      font-weight: 600;
      font-size: inherit;
      padding: 0;
      cursor: pointer;
    }}
    .summary-toggle:hover {{ text-decoration: underline; }}
    .summary-full {{ margin: 0 0 12px; }}
    footer.page {{
      text-align: center;
      padding: 48px 0 24px;
      color: var(--muted);
      font-size: 0.9rem;
    }}
    @media (min-width: 900px) {{
      body {{ padding: 60px 24px; }}
      header.hero h1 {{ font-size: 3rem; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header class="hero">
      <h1>🧠 Viv Mind</h1>
      <p>A curated watercolor journal of links worth remembering.</p>
    </header>
    <main class="feed">
{cards_html}
    </main>
    <footer class="page">
      Made with curiosity · <a href="https://github.com/vivmuk/viv-mind">vivmuk/viv-mind</a>
    </footer>
  </div>
<div id="lightbox" class="lightbox" onclick="closeLightbox()">
    <button class="lightbox-close" type="button" aria-label="Close">&times;</button>
    <img id="lightbox-img" src="" alt="" />
    <div id="lightbox-caption" class="lightbox-caption"></div>
  </div>
  <script>
    function openLightbox(src, caption) {{
      const box = document.getElementById('lightbox');
      const img = document.getElementById('lightbox-img');
      const cap = document.getElementById('lightbox-caption');
      img.src = src;
      img.alt = caption;
      cap.textContent = caption;
      box.classList.add('active');
      document.body.style.overflow = 'hidden';
    }}
    function closeLightbox() {{
      const box = document.getElementById('lightbox');
      box.classList.remove('active');
      document.body.style.overflow = '';
    }}
    document.addEventListener('keydown', e => {{
      if (e.key === 'Escape') closeLightbox();
    }});
    function toggleSummary(id) {{
      const short = document.getElementById('summary-short-' + id);
      const full = document.getElementById('summary-full-' + id);
      if (!short || !full) return;
      const isShortVisible = short.style.display !== 'none';
      short.style.display = isShortVisible ? 'none' : 'block';
      full.style.display = isShortVisible ? 'block' : 'none';
    }}
    document.querySelectorAll('.carousel').forEach(carousel => {{
      const track = carousel.querySelector('.carousel-track');
      const dots = [...carousel.querySelectorAll('.carousel-dot')];
      let index = 0;
      const count = dots.length;
      if (count <= 1) return;
      function setSlide(i) {{
        index = i;
        track.style.transform = 'translateX(' + (-index * 100) + '%)';
        dots.forEach((d, j) => d.classList.toggle('active', j === index));
      }}
      setSlide(0);
      dots.forEach((d, i) => d.addEventListener('click', () => setSlide(i)));
      let startX = 0;
      carousel.addEventListener('touchstart', e => startX = e.touches[0].clientX);
      carousel.addEventListener('touchend', e => {{
        const endX = e.changedTouches[0].clientX;
        if (startX - endX > 40) setSlide((index + 1) % count);
        if (endX - startX > 40) setSlide((index - 1 + count) % count);
      }});
    }});
  </script>
</body>
</html>
"""


def main() -> int:
    parser = argparse.ArgumentParser(description="Render viv-mind HTML feed.")
    parser.add_argument("--journal", default=str(DEFAULT_JOURNAL), help="Path to journal xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Path to output index.html")
    parser.add_argument("--limit", type=int, default=200, help="Max entries to render")
    args = parser.parse_args()

    journal_path = Path(args.journal).expanduser()
    output_path = Path(args.output).expanduser()

    entries = load_entries(journal_path)[-args.limit:]
    html = build_feed(entries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    print(f"Feed written: {output_path} ({len(entries)} entries)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
