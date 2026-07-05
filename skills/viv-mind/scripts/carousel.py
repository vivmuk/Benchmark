#!/usr/bin/env python3
"""Generate a multi-image watercolor carousel for a journal entry in parallel."""

import argparse
import base64
import json
import os
import random
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

API_BASE = "https://api.venice.ai/api/v1"
USER_AGENT = "VivMindSkill/1.0"
DEFAULT_MODEL = "gpt-image-2"

DEFAULT_REPO = Path.home() / ".openclaw" / "workspace" / "viv-mind-scaffold"
DEFAULT_ASSETS = DEFAULT_REPO / "docs" / "assets" / "infographics"
DEFAULT_JOURNAL = DEFAULT_REPO / "data" / "viv-mind-journal.xlsx"


def get_api_key() -> str | None:
    key = os.environ.get("VENICE_API_KEY", "").strip()
    if key:
        return key
    clawdbot = Path.home() / ".clawdbot" / "clawdbot.json"
    if clawdbot.exists():
        try:
            cfg = json.loads(clawdbot.read_text(encoding="utf-8"))
            for skill_name in ("venice-ai", "venice-ai-media", "viv-mind"):
                k = (
                    cfg.get("skills", {})
                    .get("entries", {})
                    .get(skill_name, {})
                    .get("env", {})
                    .get("VENICE_API_KEY", "")
                )
                if k:
                    return k.strip()
        except (json.JSONDecodeError, OSError):
            pass
    return None


def require_api_key() -> str:
    key = get_api_key()
    if not key:
        print("Error: VENICE_API_KEY not set", file=sys.stderr)
        sys.exit(2)
    return key


def load_entry(journal_path: Path, entry_id: str) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required")

    wb = openpyxl.load_workbook(journal_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("id") + 1
    except ValueError:
        raise RuntimeError("Journal missing id column")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == entry_id:
            return {h: (ws.cell(row=row, column=i + 1).value or "")
                    for i, h in enumerate(headers)}
    raise RuntimeError(f"Entry not found: {entry_id}")


def fetch_rate_limits(api_key: str) -> dict[str, dict[str, int]]:
    """Return {model_id: {rpm: int, tpm: int}} from Venice rate-limits endpoint."""
    req = urllib.request.Request(
        f"{API_BASE}/api_keys/rate_limits",
        method="GET",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Accept": "application/json",
            "User-Agent": USER_AGENT,
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Venice rate-limits error ({e.code}): {error_body}") from e

    result: dict[str, dict[str, int]] = {}
    for entry in data.get("data", {}).get("rateLimits", []):
        model_id = entry.get("apiModelId", "")
        limits = {lim["type"]: lim["amount"] for lim in entry.get("rateLimits", [])
                  if "type" in lim and "amount" in lim}
        result[model_id] = {
            "rpm": limits.get("RPM", 0),
            "tpm": limits.get("TPM", 0),
        }
    return result


def get_model_rpm(rate_limits: dict[str, dict[str, int]], model: str) -> int:
    return rate_limits.get(model, {}).get("rpm", 50)


def slugify(text: str, max_len: int = 40) -> str:
    import re
    text = str(text).lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return (text or "image")[:max_len]


def generate_image(
    api_key: str,
    prompt: str,
    width: int = 1024,
    height: int = 1280,
    model: str = DEFAULT_MODEL,
    timeout: int = 300,
    max_retries: int = 3,
) -> bytes:
    url = f"{API_BASE}/image/generate"
    payload = {
        "model": model,
        "prompt": prompt,
        "width": width,
        "height": height,
        "format": "png",
        "safe_mode": False,
        "hide_watermark": True,
    }
    body = json.dumps(payload).encode("utf-8")

    for attempt in range(max_retries + 1):
        req = urllib.request.Request(
            url,
            method="POST",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "User-Agent": USER_AGENT,
            },
            data=body,
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            images = data.get("images", [])
            if not images:
                raise RuntimeError("No image returned from Venice API")
            return base64.b64decode(images[0])
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8", errors="replace")
            # Retry on rate-limit or transient provider unavailability.
            if e.code in (429, 503) and attempt < max_retries:
                sleep_s = (2 ** attempt) + random.random()
                print(f"  HTTP {e.code}, retrying in {sleep_s:.1f}s...")
                time.sleep(sleep_s)
                continue
            raise RuntimeError(f"Venice API error ({e.code}): {error_body}") from e


def generate_slide(
    api_key: str,
    slide: dict,
    assets_dir: Path,
    entry_id: str,
    title_slug: str,
    model: str,
) -> Path:
    num = slide["number"]
    prompt = slide["prompt"]
    print(f"Generating carousel slide {num}...")
    image_bytes = generate_image(api_key, prompt, 1024, 1280, model)
    filename = f"{entry_id}-{title_slug}-{num}.png"
    image_path = assets_dir / filename
    image_path.write_bytes(image_bytes)
    print(f"  Saved slide {num}: {image_path} ({len(image_bytes) // 1024}KB)")
    return image_path


def update_entry_carousel(journal_path: Path, entry_id: str, paths: list[Path]) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    wb = openpyxl.load_workbook(journal_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("id") + 1
        img_col = headers.index("image_path") + 1
    except ValueError:
        return
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == entry_id:
            ws.cell(row=row, column=img_col, value="; ".join(str(p) for p in paths))
            wb.save(journal_path)
            return


def generate_carousel(
    plan: dict,
    entry_id: str,
    assets_dir: Path,
    journal_path: Path,
    model: str,
    workers: int = 5,
) -> list[Path]:
    api_key = require_api_key()
    slides = plan.get("slides", [])
    if not slides:
        raise RuntimeError("No slides found in carousel plan")

    assets_dir.mkdir(parents=True, exist_ok=True)
    title_slug = slugify(plan.get("title", entry_id))

    # Rate-limit awareness: clamp workers to the model's RPM.
    try:
        rate_limits = fetch_rate_limits(api_key)
        rpm = get_model_rpm(rate_limits, model)
        print(f"Venice rate limit for {model}: {rpm} RPM")
        workers = max(1, min(workers, rpm))
        print(f"Using {workers} parallel worker(s)")
    except Exception as e:
        print(f"Could not fetch rate limits: {e}; defaulting to 1 worker", file=sys.stderr)
        workers = 1

    saved_paths: list[Path] = []
    if workers == 1:
        for slide in slides:
            saved_paths.append(generate_slide(api_key, slide, assets_dir, entry_id, title_slug, model))
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(
                    generate_slide, api_key, slide, assets_dir, entry_id, title_slug, model
                ): slide for slide in slides
            }
            for future in as_completed(futures):
                saved_paths.append(future.result())

    # Sort by slide number so the journal order matches the plan.
    saved_paths.sort(key=lambda p: int(p.stem.split("-")[-1]))

    update_entry_carousel(journal_path, entry_id, saved_paths)
    print(f"Updated journal carousel paths for {entry_id}")
    return saved_paths


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a watercolor carousel for a journal entry.")
    parser.add_argument("--id", required=True, help="Journal entry id")
    parser.add_argument("--plan", required=True, help="Path to carousel plan JSON")
    parser.add_argument("--journal", default=str(DEFAULT_JOURNAL), help="Path to journal xlsx")
    parser.add_argument("--assets", default=str(DEFAULT_ASSETS), help="Output directory")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"Venice image model (default: {DEFAULT_MODEL})")
    parser.add_argument("--start", type=int, default=1, help="Slide number to start from (default: 1)")
    parser.add_argument("--workers", type=int, default=5, help="Max parallel workers (default: 5)")
    args = parser.parse_args()

    plan_path = Path(args.plan).expanduser()
    full_plan = json.loads(plan_path.read_text(encoding="utf-8"))
    slides = full_plan.get("slides", [])
    start_index = max(1, args.start)
    full_plan["slides"] = [s for s in slides if s.get("number", 0) >= start_index]

    saved_paths = generate_carousel(
        full_plan,
        args.id,
        Path(args.assets).expanduser(),
        Path(args.journal).expanduser(),
        args.model,
        workers=args.workers,
    )
    print(f"\nGenerated {len(saved_paths)} carousel images")
    for p in saved_paths:
        print(p)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
