#!/usr/bin/env python3
"""Append or update entries in viv-mind-journal.xlsx."""

import argparse
import json
import os
import sys
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    openpyxl = None  # type: ignore
    Workbook = None  # type: ignore


COLUMNS = [
    "id",
    "timestamp",
    "source_url",
    "source_type",
    "title",
    "author",
    "published_date",
    "summary",
    "key_quotes",
    "tags",
    "image_path",
    "status",
]

DEFAULT_REPO = Path.home() / ".openclaw" / "workspace" / "viv-mind-scaffold"


def _ensure_workbook(path: Path) -> Any:
    """Create or load workbook, return (wb, ws)."""
    if openpyxl is None:
        raise RuntimeError("openpyxl is required. Install: pip3 install openpyxl")
    if path.exists() and path.stat().st_size > 0:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception:
            # Corrupt/placeholder file: create fresh workbook and overwrite.
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal"
            for idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            return wb, ws
        ws = wb.active
        # Ensure all expected columns exist
        existing = [cell.value for cell in ws[1]]
        for col in COLUMNS:
            if col not in existing:
                existing.append(col)
        # Rewrite header if changed
        if existing != [cell.value for cell in ws[1]]:
            ws.delete_rows(1)
            ws.insert_rows(1)
            for idx, col_name in enumerate(existing, start=1):
                ws.cell(row=1, column=idx, value=col_name)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    for idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
    return wb, ws


def _to_cell_value(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return "" if value is None else str(value)


def _row_to_entry(ws, row_idx: int) -> dict[str, str]:
    headers = [cell.value for cell in ws[1]]
    values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
    return {h: (v if v is not None else "") for h, v in zip(headers, values)}


def entry_exists(ws, entry_id: str) -> tuple[bool, int]:
    """Check if entry id already exists, return (exists, row_index)."""
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("id") + 1
    except ValueError:
        return False, -1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=id_col).value == entry_id:
            return True, row_idx
    return False, -1


def add_entry(path: Path, data: dict[str, Any]) -> dict[str, Any]:
    """Add or update an entry in the journal."""
    wb, ws = _ensure_workbook(path)
    entry_id = data.get("id") or _new_id()
    data["id"] = entry_id
    data.setdefault("timestamp", datetime.now(timezone.utc).isoformat(timespec="seconds"))
    data.setdefault("status", "saved")

    exists, row_idx = entry_exists(ws, entry_id)
    headers = [cell.value for cell in ws[1]]

    if exists and row_idx > 0:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_to_cell_value(data.get(header, "")))
        operation = "updated"
    else:
        row_idx = ws.max_row + 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_to_cell_value(data.get(header, "")))
        operation = "added"

    # Auto-adjust column widths for readability
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(1, ws.max_row + 1):
            val = str(ws.cell(row=r, column=col_idx).value or "")
            max_len = max(max_len, min(len(val), 80))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {"operation": operation, "id": entry_id, "row": row_idx, "path": str(path)}


def _new_id() -> str:
    import uuid
    return uuid.uuid4().hex[:12]


def load_entry(path: Path, entry_id: str) -> dict[str, str] | None:
    """Load a single entry by id."""
    if not path.exists():
        return None
    wb, ws = _ensure_workbook(path)
    exists, row_idx = entry_exists(ws, entry_id)
    if exists:
        return _row_to_entry(ws, row_idx)
    return None


def list_entries(path: Path, limit: int = 100) -> list[dict[str, str]]:
    """Return recent journal entries, newest last."""
    if not path.exists():
        return []
    wb, ws = _ensure_workbook(path)
    entries = []
    for row_idx in range(2, ws.max_row + 1):
        entries.append(_row_to_entry(ws, row_idx))
    return entries[-limit:]


def parse_input(text: str) -> dict[str, Any]:
    """Parse JSON from string or file path."""
    text = text.strip()
    if not text:
        return {}
    if text == "-":
        text = sys.stdin.read()
    elif Path(text).expanduser().exists():
        text = Path(text).expanduser().read_text(encoding="utf-8")
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON input: {exc}") from exc


def main() -> int:
    parser = argparse.ArgumentParser(description="Manage viv-mind journal.")
    parser.add_argument(
        "--journal",
        default=str(DEFAULT_REPO / "data" / "viv-mind-journal.xlsx"),
        help="Path to journal xlsx",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    add_cmd = sub.add_parser("add", help="Add or update an entry")
    add_cmd.add_argument("input", nargs="?", default="-", help="JSON string, file path, or - for stdin (default)")

    get_cmd = sub.add_parser("get", help="Get entry by id")
    get_cmd.add_argument("id", help="Entry id")

    list_cmd = sub.add_parser("list", help="List recent entries")
    list_cmd.add_argument("--limit", type=int, default=100)

    args = parser.parse_args()

    journal_path = Path(args.journal).expanduser()

    if args.command == "add":
        data = parse_input(args.input)
        if not data:
            print("Error: empty input", file=sys.stderr)
            return 1
        result = add_entry(journal_path, data)
        print(json.dumps(result, indent=2))
        return 0

    if args.command == "get":
        entry = load_entry(journal_path, args.id)
        if not entry:
            print(f"Error: entry not found: {args.id}", file=sys.stderr)
            return 1
        print(json.dumps(entry, indent=2))
        return 0

    if args.command == "list":
        entries = list_entries(journal_path, limit=args.limit)
        print(json.dumps(entries, indent=2))
        return 0

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
